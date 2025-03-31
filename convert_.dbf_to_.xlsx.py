from dbfread import DBF
import openpyxl
from openpyxl.utils import get_column_letter
import re
import unicodedata
import tkinter as tk
from tkinter import filedialog

def limpiar_valor(valor):
    """Limpia caracteres no válidos para Excel."""
    if isinstance(valor, str):
        valor = unicodedata.normalize('NFKC', valor)
        valor = ''.join(c for c in valor if unicodedata.category(c) != 'Cc')
        return re.sub(r'[^\x00-\x7F]+', '', valor)
    return valor

def dbf_a_xlsx(archivo_dbf, archivo_xlsx, encoding='latin-1', archivo_tipo='xlsx'):
    """Convierte un archivo DBF a XLSX y autoajusta el ancho de las columnas."""
    tabla = DBF(archivo_dbf, encoding=encoding)
    libro_xlsx = openpyxl.Workbook()
    hoja_xlsx = libro_xlsx.active
    for i, campo in enumerate(tabla.fields):
        hoja_xlsx.cell(row=1, column=i + 1, value=campo.name)
    for fila_idx, fila in enumerate(tabla):
        for col_idx, valor in enumerate(fila.values()):
            hoja_xlsx.cell(row=fila_idx + 2, column=col_idx + 1, value=limpiar_valor(valor))
    for columna in hoja_xlsx.columns:
        max_longitud = 0
        for celda in columna:
            try:
                if len(str(celda.value)) > max_longitud:
                    max_longitud = len(str(celda.value))
            except TypeError:
                pass
        columna_ajustada = (max_longitud + 2) * 1.1 
        hoja_xlsx.column_dimensions[get_column_letter(columna[0].column)].width = columna_ajustada
    libro_xlsx.save(archivo_xlsx)

def seleccionar_archivo_dbf():
    root = tk.Tk()
    root.withdraw()  # Oculta la ventana principal de tkinter
    archivo_dbf = filedialog.askopenfilename(title="Seleccionar archivo DBF", filetypes=[("Archivos DBF", "*.dbf")])
    return archivo_dbf

def obtener_nombre_archivo_xlsx():
    root = tk.Tk()
    root.withdraw()
    nombre_archivo = filedialog.asksaveasfilename(title="Guardar archivo XLSX como", defaultextension=".xlsx", filetypes=[("Archivos XLSX", "*.xlsx")])
    return nombre_archivo

# Interfaz de usuario para seleccionar archivos y carpetas
archivo_dbf = seleccionar_archivo_dbf()
if not archivo_dbf:
    print("No se seleccionó ningún archivo DBF.")
    exit()

nombre_archivo_xlsx = obtener_nombre_archivo_xlsx()
if not nombre_archivo_xlsx:
    print("No se especificó el nombre del archivo XLSX.")
    exit()

# Intentar guardar como XLSX "estricto" (si es compatible)
try:
    dbf_a_xlsx(archivo_dbf, nombre_archivo_xlsx, archivo_tipo='xlsx')
except Exception as e:
    print(f"Error al guardar como XLSX estricto: {e}")
    # Si falla, guardar como XLSX normal
    dbf_a_xlsx(archivo_dbf, nombre_archivo_xlsx)

print("Terminado")